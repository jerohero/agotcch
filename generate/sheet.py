import openpyxl
from openpyxl.styles import PatternFill

def export_to_excel(characters: dict, file_name: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Canon Children"

    headers = ["ID", "Name", "Birth", "Father ID", "Mother ID", "Real Father ID"]
    sheet.append(headers)

    father_header_fill = PatternFill(start_color="81D41A", end_color="81D41A", fill_type="solid")
    father_row_fill = PatternFill(start_color="E8F2A1", end_color="E8F2A1", fill_type="solid")
    child_header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    child_row_fill = PatternFill(start_color="FFDBB6", end_color="FFDBB6", fill_type="solid")
    mother_header_fill = PatternFill(start_color="729FCF", end_color="729FCF", fill_type="solid")
    mother_row_fill = PatternFill(start_color="DEE6EF", end_color="DEE6EFA1", fill_type="solid")
    real_father_header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    real_father_row_fill = PatternFill(start_color="FFFFD7", end_color="FFFFD7", fill_type="solid")

    sheet[1][0].fill = child_header_fill
    sheet[1][1].fill = child_header_fill
    sheet[1][2].fill = child_header_fill
    sheet[1][3].fill = father_header_fill
    sheet[1][4].fill = mother_header_fill
    sheet[1][5].fill = real_father_header_fill

    for child_id, child in characters.items():
        if child["skipped"]:
            continue

        row = [
            child_id,
            child["name"]["primary"],
            child["birth"],
            child["father"],
            # fathers[child["father"]],
            child["mother"],
            # mothers[child["mother"]],
            child["real_father"],
            # real_fathers[child["real_father"]] if child["real_father"] else "",
        ]
        sheet.append(row)

        sheet[f"A{sheet.max_row}"].fill = child_row_fill
        sheet[f"B{sheet.max_row}"].fill = child_row_fill
        sheet[f"C{sheet.max_row}"].fill = child_row_fill
        sheet[f"D{sheet.max_row}"].fill = father_row_fill
        sheet[f"E{sheet.max_row}"].fill = mother_row_fill
        sheet[f"F{sheet.max_row}"].fill = real_father_row_fill

    # Set column widths
    # column_widths = {
    #     "A": 5,  # ID
    #     "B": 5,  # Name
    #     "C": 5, # Birth
    #     "D": 5,  # Father ID
    #     # "E": 1.5,  # Father Name
    #     "E": 5,  # Mother ID
    #     # "G": 1.5,  # Mother Name
    #     "F": 5,  # Real Father ID
    #     # "I": 1.5,  # Real Father Name
    # }

    # for col, width in column_widths.items():
    #     sheet.column_dimensions[col].width = width

    workbook.save(f"{file_name}.xlsx")
